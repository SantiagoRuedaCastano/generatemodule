import openpyxl

class ExcelHelper:

    def getWorkBook(self, path):
        return openpyxl.load_workbook(path)

    def getSheet(self, workbook, sheet):
        return workbook[sheet]

    def getRowsRange(self, sheet, fromCell, toCell):
        row = sheet[fromCell].row
        row_final = sheet[toCell].row
        col = sheet[fromCell].column
        alias = list()
        while row <= row_final:
            cell = sheet.cell(row, col)
            if not cell.value == None:
                alias.append(cell.value.strip())
            row += 1
        return alias

    def convertColsToDict(self, sheet, fromKey, toKey, colValues):
        initialrow = sheet[fromKey].row
        finalrow = sheet[toKey].row
        joinDict = {}

        while initialrow <= finalrow:
            initialcell = sheet.cell(row=initialrow, column=sheet[fromKey].column).value
            sources = sheet.cell(row=initialrow, column=sheet[colValues].column).value
            if (not initialcell == None) and (not sources == None):
                joinDict.setdefault(initialcell.replace("\n", ""), sources.strip())
            initialrow += 1
        return joinDict
